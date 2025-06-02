#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import sys
import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union
import openpyxl

# 添加项目根目录到Python路径
root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, root_dir)

from db.clickhouse_db import get_clickhouse_db, get_default_config
from enums.kline_period import KlinePeriod
from utils.logger import get_logger
from utils.path_utils import get_result_dir
from indicators.factory import IndicatorFactory

# 获取日志记录器
logger = get_logger(__name__)

class BuyPointDimensionAnalyzer:
    """
    买点维度分析器 - 对买点进行多维度分析
    
    支持对形态、趋势、时间特征等维度进行买点分析，发现买点共性特征
    """
    
    def __init__(self):
        """初始化买点维度分析器"""
        logger.info("初始化买点维度分析器")
        
        # 获取数据库连接
        config = get_default_config()
        self.ch_db = get_clickhouse_db(config=config)
        
        # 创建指标工厂
        self.indicator_factory = IndicatorFactory()
        
        # 存储分析结果
        self.analysis_results = {
            "pattern_analysis": {},
            "trend_analysis": {},
            "volume_analysis": {},
            "time_analysis": {},
            "indicator_analysis": {},
            "multi_period_analysis": {}
        }
        
        # 结果输出目录
        self.result_dir = get_result_dir()
        os.makedirs(self.result_dir, exist_ok=True)
        
        logger.info("买点维度分析器初始化完成") 

    def analyze_pattern_features(self, stock_codes: List[str], 
                               start_date: str, end_date: str,
                               period: str = 'DAILY') -> Dict[str, Any]:
        """
        分析买点K线形态特征
        
        Args:
            stock_codes: 股票代码列表
            start_date: 开始日期，格式YYYYMMDD
            end_date: 结束日期，格式YYYYMMDD
            period: K线周期
            
        Returns:
            Dict: 形态特征分析结果
        """
        logger.info(f"开始分析买点K线形态特征: {len(stock_codes)}只股票, 周期={period}")
        
        try:
            # 形态特征统计
            pattern_stats = {
                "single_candle": {},   # 单根K线形态
                "combined_candle": {}, # 组合K线形态
                "complex_pattern": {}  # 复杂形态
            }
            
            # 获取K线数据
            for stock_code in stock_codes:
                # 获取股票K线数据
                sql = f"""
                SELECT date, open, high, low, close, volume, amount
                FROM stock_{period.lower()}
                WHERE stock_code = '{stock_code}'
                  AND date >= '{start_date}' AND date <= '{end_date}'
                ORDER BY date
                """
                kline_data = self.ch_db.query_df(sql)
                
                if kline_data.empty:
                    logger.warning(f"未找到股票 {stock_code} 的K线数据")
                    continue
                
                # 分析单根K线形态
                self._analyze_single_candle_patterns(kline_data, pattern_stats["single_candle"])
                
                # 分析组合K线形态
                self._analyze_combined_candle_patterns(kline_data, pattern_stats["combined_candle"])
                
                # 分析复杂形态
                self._analyze_complex_patterns(kline_data, pattern_stats["complex_pattern"])
            
            # 计算形态出现频率
            result = {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "stock_count": len(stock_codes),
                "pattern_stats": self._calculate_pattern_frequency(pattern_stats, len(stock_codes))
            }
            
            # 识别高频特征
            result["common_features"] = self._identify_common_pattern_features(result["pattern_stats"])
            
            # 保存到分析结果
            self.analysis_results["pattern_analysis"] = result
            
            logger.info(f"买点K线形态特征分析完成")
            return result
            
        except Exception as e:
            logger.error(f"分析买点K线形态特征时出错: {e}")
            return {}
    
    def _analyze_single_candle_patterns(self, kline_data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析单根K线形态"""
        # 计算K线实体和影线
        kline_data['body'] = abs(kline_data['close'] - kline_data['open'])
        kline_data['upper_shadow'] = kline_data['high'] - kline_data[['open', 'close']].max(axis=1)
        kline_data['lower_shadow'] = kline_data[['open', 'close']].min(axis=1) - kline_data['low']
        
        # 计算前一日K线
        kline_data['prev_close'] = kline_data['close'].shift(1)
        kline_data['prev_body'] = kline_data['body'].shift(1)
        
        # 统计各种单K线形态
        # 十字星
        doji = kline_data[kline_data['body'] < kline_data['body'].mean() * 0.3]
        self._update_pattern_count(stats, "十字星", len(doji))
        
        # 锤子线
        hammer = kline_data[(kline_data['lower_shadow'] > kline_data['body'] * 2) & 
                          (kline_data['upper_shadow'] < kline_data['body'] * 0.5)]
        self._update_pattern_count(stats, "锤子线", len(hammer))
        
        # 吊颈线
        hanging_man = kline_data[(kline_data['lower_shadow'] > kline_data['body'] * 2) & 
                               (kline_data['upper_shadow'] < kline_data['body'] * 0.5) &
                               (kline_data['close'] < kline_data['open'])]
        self._update_pattern_count(stats, "吊颈线", len(hanging_man))
        
        # 长腿十字星
        long_legged_doji = kline_data[(kline_data['body'] < kline_data['body'].mean() * 0.3) &
                                    (kline_data['upper_shadow'] > kline_data['body']) &
                                    (kline_data['lower_shadow'] > kline_data['body'])]
        self._update_pattern_count(stats, "长腿十字星", len(long_legged_doji))
        
        # 射击之星
        shooting_star = kline_data[(kline_data['upper_shadow'] > kline_data['body'] * 2) & 
                                 (kline_data['lower_shadow'] < kline_data['body'] * 0.5)]
        self._update_pattern_count(stats, "射击之星", len(shooting_star))
        
        # 穿刺线
        piercing = kline_data[(kline_data['close'] > kline_data['open']) & 
                            (kline_data['open'] < kline_data['prev_close']) &
                            (kline_data['close'] > (kline_data['prev_close'] + kline_data['prev_close'].shift(1)) / 2)]
        self._update_pattern_count(stats, "穿刺线", len(piercing))
    
    def _analyze_combined_candle_patterns(self, kline_data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析组合K线形态"""
        # 获取必要的前置数据
        kline_data['prev_close'] = kline_data['close'].shift(1)
        kline_data['prev_open'] = kline_data['open'].shift(1)
        kline_data['prev_high'] = kline_data['high'].shift(1)
        kline_data['prev_low'] = kline_data['low'].shift(1)
        
        kline_data['prev2_close'] = kline_data['close'].shift(2)
        kline_data['prev2_open'] = kline_data['open'].shift(2)
        kline_data['prev2_high'] = kline_data['high'].shift(2)
        kline_data['prev2_low'] = kline_data['low'].shift(2)
        
        # 统计各种组合K线形态
        # 看涨吞没
        bullish_engulfing = kline_data[(kline_data['open'] < kline_data['prev_close']) & 
                                     (kline_data['close'] > kline_data['prev_open']) &
                                     (kline_data['prev_close'] < kline_data['prev_open'])]
        self._update_pattern_count(stats, "看涨吞没", len(bullish_engulfing))
        
        # 看跌吞没
        bearish_engulfing = kline_data[(kline_data['open'] > kline_data['prev_close']) & 
                                     (kline_data['close'] < kline_data['prev_open']) &
                                     (kline_data['prev_close'] > kline_data['prev_open'])]
        self._update_pattern_count(stats, "看跌吞没", len(bearish_engulfing))
        
        # 启明星
        morning_star = kline_data[(kline_data['prev2_close'] < kline_data['prev2_open']) &  # 第一天下跌
                                (abs(kline_data['prev_close'] - kline_data['prev_open']) < 
                                 abs(kline_data['prev2_close'] - kline_data['prev2_open']) * 0.3) &  # 第二天十字星
                                (kline_data['close'] > kline_data['open']) &  # 第三天上涨
                                (kline_data['close'] > (kline_data['prev2_open'] + kline_data['prev2_close']) / 2)]  # 收复部分跌幅
        self._update_pattern_count(stats, "启明星", len(morning_star))
        
        # 黄昏星
        evening_star = kline_data[(kline_data['prev2_close'] > kline_data['prev2_open']) &  # 第一天上涨
                                (abs(kline_data['prev_close'] - kline_data['prev_open']) < 
                                 abs(kline_data['prev2_close'] - kline_data['prev2_open']) * 0.3) &  # 第二天十字星
                                (kline_data['close'] < kline_data['open']) &  # 第三天下跌
                                (kline_data['close'] < (kline_data['prev2_open'] + kline_data['prev2_close']) / 2)]  # 回撤部分涨幅
        self._update_pattern_count(stats, "黄昏星", len(evening_star))
        
        # 三只乌鸦
        three_black_crows = []
        for i in range(3, len(kline_data)):
            if (kline_data.iloc[i-3]['close'] > kline_data.iloc[i-3]['open'] and
                kline_data.iloc[i-2]['close'] < kline_data.iloc[i-2]['open'] and
                kline_data.iloc[i-1]['close'] < kline_data.iloc[i-1]['open'] and
                kline_data.iloc[i]['close'] < kline_data.iloc[i]['open'] and
                kline_data.iloc[i-1]['close'] < kline_data.iloc[i-2]['close'] and
                kline_data.iloc[i]['close'] < kline_data.iloc[i-1]['close']):
                three_black_crows.append(i)
        self._update_pattern_count(stats, "三只乌鸦", len(three_black_crows))
        
        # 三白兵
        three_white_soldiers = []
        for i in range(3, len(kline_data)):
            if (kline_data.iloc[i-3]['close'] < kline_data.iloc[i-3]['open'] and
                kline_data.iloc[i-2]['close'] > kline_data.iloc[i-2]['open'] and
                kline_data.iloc[i-1]['close'] > kline_data.iloc[i-1]['open'] and
                kline_data.iloc[i]['close'] > kline_data.iloc[i]['open'] and
                kline_data.iloc[i-1]['close'] > kline_data.iloc[i-2]['close'] and
                kline_data.iloc[i]['close'] > kline_data.iloc[i-1]['close']):
                three_white_soldiers.append(i)
        self._update_pattern_count(stats, "三白兵", len(three_white_soldiers))
    
    def _analyze_complex_patterns(self, kline_data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析复杂形态"""
        # 移动平均线
        kline_data['ma5'] = kline_data['close'].rolling(5).mean()
        kline_data['ma10'] = kline_data['close'].rolling(10).mean()
        kline_data['ma20'] = kline_data['close'].rolling(20).mean()
        
        # 双底形态检测
        double_bottom = self._detect_double_bottom(kline_data)
        self._update_pattern_count(stats, "双底", len(double_bottom))
        
        # 双顶形态检测
        double_top = self._detect_double_top(kline_data)
        self._update_pattern_count(stats, "双顶", len(double_top))
        
        # 头肩底形态检测
        head_and_shoulders_bottom = self._detect_head_and_shoulders_bottom(kline_data)
        self._update_pattern_count(stats, "头肩底", len(head_and_shoulders_bottom))
        
        # 头肩顶形态检测
        head_and_shoulders_top = self._detect_head_and_shoulders_top(kline_data)
        self._update_pattern_count(stats, "头肩顶", len(head_and_shoulders_top))
        
        # 三角形整理形态
        triangle = self._detect_triangle(kline_data)
        self._update_pattern_count(stats, "三角形整理", len(triangle))
        
        # 旗形整理
        flag = self._detect_flag(kline_data)
        self._update_pattern_count(stats, "旗形整理", len(flag))
        
        # 杯柄形态
        cup_and_handle = self._detect_cup_and_handle(kline_data)
        self._update_pattern_count(stats, "杯柄形态", len(cup_and_handle))
    
    def _detect_double_bottom(self, data: pd.DataFrame) -> List[int]:
        """检测双底形态"""
        # 简化检测逻辑
        result = []
        window = 20
        
        for i in range(window, len(data) - window):
            # 获取窗口数据
            window_data = data.iloc[i-window:i+window]
            
            # 找到窗口内的最低点
            min_indices = window_data.loc[window_data['low'] == window_data['low'].min()].index
            
            if len(min_indices) < 2:
                continue
                
            # 检查两个最低点之间的距离和价格差异
            min_idx1 = min_indices[0]
            min_idx2 = min_indices[-1]
            
            if min_idx2 - min_idx1 < 10:
                continue
                
            min_price1 = data.loc[min_idx1, 'low']
            min_price2 = data.loc[min_idx2, 'low']
            
            # 底部高度差异不大
            if abs(min_price2 - min_price1) / min_price1 > 0.05:
                continue
                
            # 检查中间是否有明显的反弹
            mid_idx = (min_idx1 + min_idx2) // 2
            mid_price = data.loc[mid_idx:mid_idx+5, 'high'].max()
            
            if (mid_price - min_price1) / min_price1 < 0.03:
                continue
                
            result.append(i)
            
        return result
    
    def _detect_double_top(self, data: pd.DataFrame) -> List[int]:
        """检测双顶形态"""
        # 简化检测逻辑
        result = []
        window = 20
        
        for i in range(window, len(data) - window):
            # 获取窗口数据
            window_data = data.iloc[i-window:i+window]
            
            # 找到窗口内的最高点
            max_indices = window_data.loc[window_data['high'] == window_data['high'].max()].index
            
            if len(max_indices) < 2:
                continue
                
            # 检查两个最高点之间的距离和价格差异
            max_idx1 = max_indices[0]
            max_idx2 = max_indices[-1]
            
            if max_idx2 - max_idx1 < 10:
                continue
                
            max_price1 = data.loc[max_idx1, 'high']
            max_price2 = data.loc[max_idx2, 'high']
            
            # 顶部高度差异不大
            if abs(max_price2 - max_price1) / max_price1 > 0.05:
                continue
                
            # 检查中间是否有明显的回调
            mid_idx = (max_idx1 + max_idx2) // 2
            mid_price = data.loc[mid_idx:mid_idx+5, 'low'].min()
            
            if (max_price1 - mid_price) / max_price1 < 0.03:
                continue
                
            result.append(i)
            
        return result
    
    def _detect_head_and_shoulders_bottom(self, data: pd.DataFrame) -> List[int]:
        """检测头肩底形态"""
        # 简化检测逻辑
        result = []
        window = 30
        
        for i in range(window, len(data) - window):
            # 待实现
            pass
            
        return result
    
    def _detect_head_and_shoulders_top(self, data: pd.DataFrame) -> List[int]:
        """检测头肩顶形态"""
        # 简化检测逻辑
        result = []
        window = 30
        
        for i in range(window, len(data) - window):
            # 待实现
            pass
            
        return result
    
    def _detect_triangle(self, data: pd.DataFrame) -> List[int]:
        """检测三角形整理形态"""
        # 简化检测逻辑
        result = []
        window = 20
        
        for i in range(window, len(data) - window):
            # 待实现
            pass
            
        return result
    
    def _detect_flag(self, data: pd.DataFrame) -> List[int]:
        """检测旗形整理形态"""
        # 简化检测逻辑
        result = []
        window = 15
        
        for i in range(window, len(data) - window):
            # 待实现
            pass
            
        return result
    
    def _detect_cup_and_handle(self, data: pd.DataFrame) -> List[int]:
        """检测杯柄形态"""
        # 简化检测逻辑
        result = []
        window = 40
        
        for i in range(window, len(data) - window):
            # 待实现
            pass
            
        return result
    
    def _update_pattern_count(self, stats: Dict[str, int], pattern_name: str, count: int) -> None:
        """更新形态计数"""
        if pattern_name in stats:
            stats[pattern_name] += count
        else:
            stats[pattern_name] = count
    
    def _calculate_pattern_frequency(self, pattern_stats: Dict[str, Dict[str, int]], 
                                    stock_count: int) -> Dict[str, Any]:
        """计算形态出现频率"""
        result = {}
        
        for category, patterns in pattern_stats.items():
            category_result = {}
            
            for pattern, count in patterns.items():
                frequency = count / stock_count if stock_count > 0 else 0
                category_result[pattern] = {
                    "count": count,
                    "frequency": frequency
                }
            
            # 按频率排序
            sorted_patterns = {k: v for k, v in sorted(
                category_result.items(), 
                key=lambda item: item[1]["frequency"], 
                reverse=True
            )}
            
            result[category] = sorted_patterns
        
        return result
    
    def _identify_common_pattern_features(self, pattern_stats: Dict[str, Any]) -> List[Dict[str, Any]]:
        """识别共性形态特征"""
        common_features = []
        
        # 设置频率阈值
        threshold = 0.2
        
        # 遍历各类形态
        for pattern_type, patterns in pattern_stats.items():
            for pattern, stats in patterns.items():
                if stats["frequency"] >= threshold:
                    common_features.append({
                        "pattern": pattern,
                        "frequency": stats["frequency"],
                        "count": stats["count"],
                        "pattern_type": pattern_type
                    })
        
        # 按频率排序
        common_features.sort(key=lambda x: x["frequency"], reverse=True)
        
        return common_features

    def analyze_trend_features(self, stock_codes: List[str], 
                            start_date: str, end_date: str,
                            period: str = 'DAILY') -> Dict[str, Any]:
        """
        分析买点趋势特征
        
        Args:
            stock_codes: 股票代码列表
            start_date: 开始日期，格式YYYYMMDD
            end_date: 结束日期，格式YYYYMMDD
            period: K线周期
            
        Returns:
            Dict: 趋势特征分析结果
        """
        logger.info(f"开始分析买点趋势特征: {len(stock_codes)}只股票, 周期={period}")
        
        try:
            # 趋势特征统计
            trend_stats = {
                "ma_trend": {},     # 均线趋势
                "price_trend": {},  # 价格趋势
                "trend_strength": {} # 趋势强度
            }
            
            # 获取K线数据
            for stock_code in stock_codes:
                # 获取股票K线数据，前推30天以计算均线
                start_date_extended = self._get_extended_date(start_date, days=30)
                
                sql = f"""
                SELECT date, open, high, low, close, volume, amount
                FROM stock_{period.lower()}
                WHERE stock_code = '{stock_code}'
                  AND date >= '{start_date_extended}' AND date <= '{end_date}'
                ORDER BY date
                """
                kline_data = self.ch_db.query_df(sql)
                
                if kline_data.empty:
                    logger.warning(f"未找到股票 {stock_code} 的K线数据")
                    continue
                
                # 计算均线
                kline_data['ma5'] = kline_data['close'].rolling(5).mean()
                kline_data['ma10'] = kline_data['close'].rolling(10).mean()
                kline_data['ma20'] = kline_data['close'].rolling(20).mean()
                kline_data['ma30'] = kline_data['close'].rolling(30).mean()
                
                # 过滤掉扩展的日期数据
                valid_data = kline_data[kline_data['date'] >= start_date]
                
                if valid_data.empty:
                    continue
                
                # 分析均线趋势
                self._analyze_ma_trend(valid_data, trend_stats["ma_trend"])
                
                # 分析价格趋势
                self._analyze_price_trend(valid_data, trend_stats["price_trend"])
                
                # 分析趋势强度
                self._analyze_trend_strength(valid_data, trend_stats["trend_strength"])
            
            # 计算趋势特征频率
            result = {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "stock_count": len(stock_codes),
                "trend_stats": self._calculate_trend_frequency(trend_stats, len(stock_codes))
            }
            
            # 识别高频特征
            result["common_features"] = self._identify_common_trend_features(result["trend_stats"])
            
            # 保存到分析结果
            self.analysis_results["trend_analysis"] = result
            
            logger.info(f"买点趋势特征分析完成")
            return result
            
        except Exception as e:
            logger.error(f"分析买点趋势特征时出错: {e}")
            return {}
    
    def _analyze_ma_trend(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析均线趋势"""
        # MA5上穿MA10
        ma5_cross_above_ma10 = data[(data['ma5'] > data['ma10']) & (data['ma5'].shift(1) <= data['ma10'].shift(1))]
        self._update_pattern_count(stats, "MA5上穿MA10", len(ma5_cross_above_ma10))
        
        # MA5下穿MA10
        ma5_cross_below_ma10 = data[(data['ma5'] < data['ma10']) & (data['ma5'].shift(1) >= data['ma10'].shift(1))]
        self._update_pattern_count(stats, "MA5下穿MA10", len(ma5_cross_below_ma10))
        
        # MA10上穿MA20
        ma10_cross_above_ma20 = data[(data['ma10'] > data['ma20']) & (data['ma10'].shift(1) <= data['ma20'].shift(1))]
        self._update_pattern_count(stats, "MA10上穿MA20", len(ma10_cross_above_ma20))
        
        # MA10下穿MA20
        ma10_cross_below_ma20 = data[(data['ma10'] < data['ma20']) & (data['ma10'].shift(1) >= data['ma20'].shift(1))]
        self._update_pattern_count(stats, "MA10下穿MA20", len(ma10_cross_below_ma20))
        
        # 均线多头排列（MA5 > MA10 > MA20 > MA30）
        bullish_alignment = data[(data['ma5'] > data['ma10']) & (data['ma10'] > data['ma20']) & (data['ma20'] > data['ma30'])]
        self._update_pattern_count(stats, "均线多头排列", len(bullish_alignment))
        
        # 均线空头排列（MA5 < MA10 < MA20 < MA30）
        bearish_alignment = data[(data['ma5'] < data['ma10']) & (data['ma10'] < data['ma20']) & (data['ma20'] < data['ma30'])]
        self._update_pattern_count(stats, "均线空头排列", len(bearish_alignment))
        
        # 均线交叉密集
        ma_crosses = 0
        for i in range(1, len(data)):
            crosses = 0
            if (data.iloc[i]['ma5'] > data.iloc[i]['ma10'] and data.iloc[i-1]['ma5'] <= data.iloc[i-1]['ma10']) or \
               (data.iloc[i]['ma5'] < data.iloc[i]['ma10'] and data.iloc[i-1]['ma5'] >= data.iloc[i-1]['ma10']):
                crosses += 1
            if (data.iloc[i]['ma10'] > data.iloc[i]['ma20'] and data.iloc[i-1]['ma10'] <= data.iloc[i-1]['ma20']) or \
               (data.iloc[i]['ma10'] < data.iloc[i]['ma20'] and data.iloc[i-1]['ma10'] >= data.iloc[i-1]['ma20']):
                crosses += 1
            if crosses >= 2:
                ma_crosses += 1
        self._update_pattern_count(stats, "均线交叉密集", ma_crosses)
        
        # 均线收敛
        ma_converge = data[((data['ma5'] - data['ma30']).abs() < (data['ma5'].shift(5) - data['ma30'].shift(5)).abs() * 0.8)]
        self._update_pattern_count(stats, "均线收敛", len(ma_converge))
        
        # 均线发散
        ma_diverge = data[((data['ma5'] - data['ma30']).abs() > (data['ma5'].shift(5) - data['ma30'].shift(5)).abs() * 1.2)]
        self._update_pattern_count(stats, "均线发散", len(ma_diverge))
    
    def _analyze_price_trend(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析价格趋势"""
        # 计算价格变动百分比
        data['price_change_pct'] = data['close'].pct_change()
        
        # 计算n日涨跌幅
        data['change_5d'] = data['close'] / data['close'].shift(5) - 1 if len(data) >= 5 else np.nan
        data['change_10d'] = data['close'] / data['close'].shift(10) - 1 if len(data) >= 10 else np.nan
        data['change_20d'] = data['close'] / data['close'].shift(20) - 1 if len(data) >= 20 else np.nan
        
        # 连续上涨
        consecutive_up = 0
        for i in range(1, len(data)):
            if data.iloc[i]['close'] > data.iloc[i-1]['close']:
                consecutive_up += 1
                if consecutive_up >= 3:  # 连续3日上涨
                    self._update_pattern_count(stats, "连续3日上涨", 1)
                if consecutive_up >= 5:  # 连续5日上涨
                    self._update_pattern_count(stats, "连续5日上涨", 1)
            else:
                consecutive_up = 0
                
        # 连续下跌
        consecutive_down = 0
        for i in range(1, len(data)):
            if data.iloc[i]['close'] < data.iloc[i-1]['close']:
                consecutive_down += 1
                if consecutive_down >= 3:  # 连续3日下跌
                    self._update_pattern_count(stats, "连续3日下跌", 1)
                if consecutive_down >= 5:  # 连续5日下跌
                    self._update_pattern_count(stats, "连续5日下跌", 1)
            else:
                consecutive_down = 0
        
        # 大涨
        big_up = data[data['price_change_pct'] > 0.05]  # 单日涨幅超过5%
        self._update_pattern_count(stats, "单日大涨(>5%)", len(big_up))
        
        # 大跌
        big_down = data[data['price_change_pct'] < -0.05]  # 单日跌幅超过5%
        self._update_pattern_count(stats, "单日大跌(>5%)", len(big_down))
        
        # 高位回调
        high_pullback = []
        window = 10
        for i in range(window, len(data)):
            # 前window天的最高价
            max_high = data.iloc[i-window:i]['high'].max()
            max_idx = data.iloc[i-window:i]['high'].idxmax()
            
            # 如果当前收盘价比最高价回调超过5%，且最高价出现在3天前以上
            if (max_high - data.iloc[i]['close']) / max_high > 0.05 and i - max_idx > 3:
                high_pullback.append(i)
        self._update_pattern_count(stats, "高位回调", len(high_pullback))
        
        # 低位反弹
        low_rebound = []
        for i in range(window, len(data)):
            # 前window天的最低价
            min_low = data.iloc[i-window:i]['low'].min()
            min_idx = data.iloc[i-window:i]['low'].idxmin()
            
            # 如果当前收盘价比最低价反弹超过5%，且最低价出现在3天前以上
            if (data.iloc[i]['close'] - min_low) / min_low > 0.05 and i - min_idx > 3:
                low_rebound.append(i)
        self._update_pattern_count(stats, "低位反弹", len(low_rebound))
        
        # 突破前高
        break_high = []
        for i in range(window, len(data)):
            prev_high = data.iloc[i-window:i-1]['high'].max()
            if data.iloc[i]['close'] > prev_high:
                break_high.append(i)
        self._update_pattern_count(stats, "突破前高", len(break_high))
        
        # 跌破前低
        break_low = []
        for i in range(window, len(data)):
            prev_low = data.iloc[i-window:i-1]['low'].min()
            if data.iloc[i]['close'] < prev_low:
                break_low.append(i)
        self._update_pattern_count(stats, "跌破前低", len(break_low))
    
    def _analyze_trend_strength(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析趋势强度"""
        # 计算ADX (Average Directional Index)
        try:
            import talib
            data['adx'] = talib.ADX(data['high'].values, data['low'].values, data['close'].values, timeperiod=14)
            
            # 强趋势 (ADX > 25)
            strong_trend = data[data['adx'] > 25]
            self._update_pattern_count(stats, "强趋势(ADX>25)", len(strong_trend))
            
            # 极强趋势 (ADX > 50)
            very_strong_trend = data[data['adx'] > 50]
            self._update_pattern_count(stats, "极强趋势(ADX>50)", len(very_strong_trend))
            
            # 无趋势 (ADX < 20)
            no_trend = data[data['adx'] < 20]
            self._update_pattern_count(stats, "无趋势(ADX<20)", len(no_trend))
            
        except ImportError:
            # 如果没有talib，使用简化的趋势判断
            logger.warning("未安装talib库，使用简化的趋势判断")
            
            # 计算简单的趋势强度：连续上涨或下跌的幅度
            up_strength = []
            down_strength = []
            
            for i in range(5, len(data)):
                # 过去5天的涨跌幅
                change_5d = data.iloc[i]['close'] / data.iloc[i-5]['close'] - 1
                
                if change_5d > 0.1:  # 5天涨幅超过10%
                    up_strength.append(i)
                elif change_5d < -0.1:  # 5天跌幅超过10%
                    down_strength.append(i)
            
            self._update_pattern_count(stats, "强上涨趋势(5日涨幅>10%)", len(up_strength))
            self._update_pattern_count(stats, "强下跌趋势(5日跌幅>10%)", len(down_strength))
    
    def _get_extended_date(self, date_str: str, days: int) -> str:
        """获取向前推days天的日期"""
        date_obj = datetime.strptime(date_str, "%Y%m%d")
        extended_date = date_obj - timedelta(days=days)
        return extended_date.strftime("%Y%m%d")
    
    def _calculate_trend_frequency(self, trend_stats: Dict[str, Dict[str, int]], 
                                  stock_count: int) -> Dict[str, Any]:
        """计算趋势特征频率"""
        return self._calculate_pattern_frequency(trend_stats, stock_count)
    
    def _identify_common_trend_features(self, trend_stats: Dict[str, Any]) -> List[Dict[str, Any]]:
        """识别共性趋势特征"""
        return self._identify_common_pattern_features(trend_stats)

    def analyze_volume_features(self, stock_codes: List[str], 
                              start_date: str, end_date: str,
                              period: str = 'DAILY') -> Dict[str, Any]:
        """
        分析买点量能特征
        
        Args:
            stock_codes: 股票代码列表
            start_date: 开始日期，格式YYYYMMDD
            end_date: 结束日期，格式YYYYMMDD
            period: K线周期
            
        Returns:
            Dict: 量能特征分析结果
        """
        logger.info(f"开始分析买点量能特征: {len(stock_codes)}只股票, 周期={period}")
        
        try:
            # 量能特征统计
            volume_stats = {
                "volume_change": {},   # 成交量变化
                "volume_pattern": {},  # 成交量形态
                "price_volume": {}     # 量价关系
            }
            
            # 获取K线数据
            for stock_code in stock_codes:
                # 获取股票K线数据，前推20天以计算均量
                start_date_extended = self._get_extended_date(start_date, days=20)
                
                sql = f"""
                SELECT date, open, high, low, close, volume, amount
                FROM stock_{period.lower()}
                WHERE stock_code = '{stock_code}'
                  AND date >= '{start_date_extended}' AND date <= '{end_date}'
                ORDER BY date
                """
                kline_data = self.ch_db.query_df(sql)
                
                if kline_data.empty:
                    logger.warning(f"未找到股票 {stock_code} 的K线数据")
                    continue
                
                # 计算均量
                kline_data['volume_ma5'] = kline_data['volume'].rolling(5).mean()
                kline_data['volume_ma10'] = kline_data['volume'].rolling(10).mean()
                kline_data['volume_ma20'] = kline_data['volume'].rolling(20).mean()
                
                # 过滤掉扩展的日期数据
                valid_data = kline_data[kline_data['date'] >= start_date]
                
                if valid_data.empty:
                    continue
                
                # 分析成交量变化
                self._analyze_volume_change(valid_data, volume_stats["volume_change"])
                
                # 分析成交量形态
                self._analyze_volume_pattern(valid_data, volume_stats["volume_pattern"])
                
                # 分析量价关系
                self._analyze_price_volume_relation(valid_data, volume_stats["price_volume"])
            
            # 计算量能特征频率
            result = {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "stock_count": len(stock_codes),
                "volume_stats": self._calculate_volume_frequency(volume_stats, len(stock_codes))
            }
            
            # 识别高频特征
            result["common_features"] = self._identify_common_volume_features(result["volume_stats"])
            
            # 保存到分析结果
            self.analysis_results["volume_analysis"] = result
            
            logger.info(f"买点量能特征分析完成")
            return result
            
        except Exception as e:
            logger.error(f"分析买点量能特征时出错: {e}")
            return {}
    
    def _analyze_volume_change(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析成交量变化"""
        # 计算量比（当日成交量/5日均量）
        data['volume_ratio'] = data['volume'] / data['volume_ma5']
        
        # 放量（成交量大于5日均量的2倍）
        volume_surge = data[data['volume_ratio'] > 2.0]
        self._update_pattern_count(stats, "放量(>2倍均量)", len(volume_surge))
        
        # 大幅放量（成交量大于5日均量的3倍）
        large_volume_surge = data[data['volume_ratio'] > 3.0]
        self._update_pattern_count(stats, "大幅放量(>3倍均量)", len(large_volume_surge))
        
        # 极度放量（成交量大于5日均量的5倍）
        extreme_volume_surge = data[data['volume_ratio'] > 5.0]
        self._update_pattern_count(stats, "极度放量(>5倍均量)", len(extreme_volume_surge))
        
        # 缩量（成交量小于5日均量的0.5倍）
        volume_shrink = data[data['volume_ratio'] < 0.5]
        self._update_pattern_count(stats, "缩量(<0.5倍均量)", len(volume_shrink))
        
        # 极度缩量（成交量小于5日均量的0.3倍）
        extreme_volume_shrink = data[data['volume_ratio'] < 0.3]
        self._update_pattern_count(stats, "极度缩量(<0.3倍均量)", len(extreme_volume_shrink))
        
        # 连续放量
        consecutive_surge = 0
        for i in range(1, len(data)):
            if data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 1.5:
                consecutive_surge += 1
                if consecutive_surge >= 3:  # 连续3日放量
                    self._update_pattern_count(stats, "连续3日放量", 1)
            else:
                consecutive_surge = 0
        
        # 连续缩量
        consecutive_shrink = 0
        for i in range(1, len(data)):
            if data.iloc[i]['volume'] < data.iloc[i]['volume_ma5'] * 0.8:
                consecutive_shrink += 1
                if consecutive_shrink >= 3:  # 连续3日缩量
                    self._update_pattern_count(stats, "连续3日缩量", 1)
            else:
                consecutive_shrink = 0
        
        # 量能突变（今日成交量比昨日成交量增加200%以上）
        volume_change = data.copy()
        volume_change['volume_change_ratio'] = volume_change['volume'] / volume_change['volume'].shift(1)
        volume_spike = volume_change[volume_change['volume_change_ratio'] > 3.0]
        self._update_pattern_count(stats, "量能突变(日环比>3倍)", len(volume_spike))
    
    def _analyze_volume_pattern(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析成交量形态"""
        # 梯量上升（连续3日以上量能逐步放大）
        volume_step_up = 0
        for i in range(3, len(data)):
            if (data.iloc[i]['volume'] > data.iloc[i-1]['volume'] > 
                data.iloc[i-2]['volume'] > data.iloc[i-3]['volume']):
                volume_step_up += 1
        self._update_pattern_count(stats, "梯量上升", volume_step_up)
        
        # 梯量下降（连续3日以上量能逐步萎缩）
        volume_step_down = 0
        for i in range(3, len(data)):
            if (data.iloc[i]['volume'] < data.iloc[i-1]['volume'] < 
                data.iloc[i-2]['volume'] < data.iloc[i-3]['volume']):
                volume_step_down += 1
        self._update_pattern_count(stats, "梯量下降", volume_step_down)
        
        # 量能分布（单日成交量在总成交量中的占比）
        if len(data) >= 5:
            for i in range(5, len(data)):
                # 计算5日内成交量总和
                total_volume = data.iloc[i-5:i]['volume'].sum()
                # 今日成交量占比
                today_ratio = data.iloc[i-1]['volume'] / total_volume if total_volume > 0 else 0
                
                # 单日占比超过40%，说明成交高度集中
                if today_ratio > 0.4:
                    self._update_pattern_count(stats, "量能高度集中(单日>40%)", 1)
        
        # 先量后价（先出现放量，后出现价格变动）
        volume_lead_price = 0
        for i in range(3, len(data)):
            # 前天放量（大于5日均量的1.5倍）但价格变动小
            if (data.iloc[i-2]['volume'] > data.iloc[i-2]['volume_ma5'] * 1.5 and
                abs(data.iloc[i-2]['close'] / data.iloc[i-3]['close'] - 1) < 0.02 and
                # 今天价格变动大
                abs(data.iloc[i]['close'] / data.iloc[i-1]['close'] - 1) > 0.03):
                volume_lead_price += 1
        self._update_pattern_count(stats, "先量后价", volume_lead_price)
        
        # 先价后量（先出现价格变动，后出现放量）
        price_lead_volume = 0
        for i in range(3, len(data)):
            # 前天价格变动大但量能一般
            if (abs(data.iloc[i-2]['close'] / data.iloc[i-3]['close'] - 1) > 0.03 and
                data.iloc[i-2]['volume'] < data.iloc[i-2]['volume_ma5'] * 1.5 and
                # 今天放量
                data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 1.5):
                price_lead_volume += 1
        self._update_pattern_count(stats, "先价后量", price_lead_volume)
    
    def _analyze_price_volume_relation(self, data: pd.DataFrame, stats: Dict[str, int]) -> None:
        """分析量价关系"""
        # 量增价增（同向）
        volume_price_up = 0
        for i in range(1, len(data)):
            if (data.iloc[i]['volume'] > data.iloc[i-1]['volume'] * 1.2 and
                data.iloc[i]['close'] > data.iloc[i-1]['close']):
                volume_price_up += 1
        self._update_pattern_count(stats, "量增价增", volume_price_up)
        
        # 量增价减（背离）
        volume_up_price_down = 0
        for i in range(1, len(data)):
            if (data.iloc[i]['volume'] > data.iloc[i-1]['volume'] * 1.2 and
                data.iloc[i]['close'] < data.iloc[i-1]['close']):
                volume_up_price_down += 1
        self._update_pattern_count(stats, "量增价减", volume_up_price_down)
        
        # 量减价增（背离）
        volume_down_price_up = 0
        for i in range(1, len(data)):
            if (data.iloc[i]['volume'] < data.iloc[i-1]['volume'] * 0.8 and
                data.iloc[i]['close'] > data.iloc[i-1]['close']):
                volume_down_price_up += 1
        self._update_pattern_count(stats, "量减价增", volume_down_price_up)
        
        # 量减价减（同向）
        volume_price_down = 0
        for i in range(1, len(data)):
            if (data.iloc[i]['volume'] < data.iloc[i-1]['volume'] * 0.8 and
                data.iloc[i]['close'] < data.iloc[i-1]['close']):
                volume_price_down += 1
        self._update_pattern_count(stats, "量减价减", volume_price_down)
        
        # 大阳大量（单日上涨超过3%且放量超过1.5倍）
        big_up_big_volume = 0
        for i in range(1, len(data)):
            if (data.iloc[i]['close'] / data.iloc[i-1]['close'] - 1 > 0.03 and
                data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 1.5):
                big_up_big_volume += 1
        self._update_pattern_count(stats, "大阳大量", big_up_big_volume)
        
        # 大阴大量（单日下跌超过3%且放量超过1.5倍）
        big_down_big_volume = 0
        for i in range(1, len(data)):
            if (data.iloc[i-1]['close'] / data.iloc[i]['close'] - 1 > 0.03 and
                data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 1.5):
                big_down_big_volume += 1
        self._update_pattern_count(stats, "大阴大量", big_down_big_volume)
        
        # 小阳大量（单日上涨小于2%但放量超过2倍）
        small_up_big_volume = 0
        for i in range(1, len(data)):
            if (0 < data.iloc[i]['close'] / data.iloc[i-1]['close'] - 1 < 0.02 and
                data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 2.0):
                small_up_big_volume += 1
        self._update_pattern_count(stats, "小阳大量", small_up_big_volume)
        
        # 小阴大量（单日下跌小于2%但放量超过2倍）
        small_down_big_volume = 0
        for i in range(1, len(data)):
            if (0 < data.iloc[i-1]['close'] / data.iloc[i]['close'] - 1 < 0.02 and
                data.iloc[i]['volume'] > data.iloc[i]['volume_ma5'] * 2.0):
                small_down_big_volume += 1
        self._update_pattern_count(stats, "小阴大量", small_down_big_volume)
    
    def _calculate_volume_frequency(self, volume_stats: Dict[str, Dict[str, int]], 
                                  stock_count: int) -> Dict[str, Any]:
        """计算量能特征频率"""
        return self._calculate_pattern_frequency(volume_stats, stock_count)
    
    def _identify_common_volume_features(self, volume_stats: Dict[str, Any]) -> List[Dict[str, Any]]:
        """识别共性量能特征"""
        return self._identify_common_pattern_features(volume_stats)

    def analyze_multi_period(self, stock_codes: List[str], date: str, 
                          periods: List[str] = ['DAILY', 'WEEKLY']) -> Dict[str, Any]:
        """
        多周期分析
        
        Args:
            stock_codes: 股票代码列表
            date: 分析日期，格式YYYYMMDD
            periods: 周期列表
            
        Returns:
            Dict: 多周期分析结果
        """
        logger.info(f"开始多周期分析: {len(stock_codes)}只股票, 周期={periods}")
        
        try:
            # 获取每个周期的特征分析
            period_results = {}
            all_features = []  # 收集所有特征用于关联分析
            
            for period in periods:
                # 获取对应周期的数据起止日期
                start_date, end_date = self._get_period_date_range(date, period)
                
                # 获取形态特征
                pattern_features = self.analyze_pattern_features(stock_codes, start_date, end_date, period)
                
                # 获取趋势特征
                trend_features = self.analyze_trend_features(stock_codes, start_date, end_date, period)
                
                # 获取量能特征
                volume_features = self.analyze_volume_features(stock_codes, start_date, end_date, period)
                
                # 合并结果
                period_results[period] = {
                    "pattern_features": pattern_features.get("common_features", []),
                    "trend_features": trend_features.get("common_features", []),
                    "volume_features": volume_features.get("common_features", [])
                }
                
                # 收集特征用于关联分析
                for feature_type, features in period_results[period].items():
                    for feature in features:
                        feature_data = feature.copy()
                        feature_data["period"] = period
                        feature_data["type"] = feature_type.replace("_features", "")
                        all_features.append(feature_data)
            
            # 寻找多周期共性特征
            common_features = self._find_multi_period_common_features(period_results)
            
            # 使用多维度特征筛选
            # 筛选高频率特征
            high_frequency_criteria = {"min_frequency": 0.3}
            high_frequency_features = self.filter_features(all_features, high_frequency_criteria)
            
            # 按特征类型筛选
            pattern_criteria = {"types": ["pattern"]}
            pattern_features = self.filter_features(all_features, pattern_criteria)
            
            trend_criteria = {"types": ["trend"]}
            trend_features = self.filter_features(all_features, trend_criteria)
            
            volume_criteria = {"types": ["volume"]}
            volume_features = self.filter_features(all_features, volume_criteria)
            
            # 添加特征股票和日期信息用于关联分析
            feature_data_for_correlation = []
            for feature in all_features:
                for stock_code in stock_codes:
                    feature_data_for_correlation.append({
                        "stock_code": stock_code,
                        "date": date,
                        "pattern": feature.get("pattern", feature.get("feature", "")),
                        "type": feature.get("type", ""),
                        "period": feature.get("period", "")
                    })
            
            # 分析特征关联性
            correlation_result = self.analyze_feature_correlation(feature_data_for_correlation)
            
            # 分析特征时序模式
            temporal_result = self.analyze_temporal_patterns(feature_data_for_correlation)
            
            # 优化特征组合
            optimized_features = self.optimize_feature_combination(
                high_frequency_features, "frequency", max_features=7)
            
            # 构建结果
            result = {
                "date": date,
                "periods": periods,
                "stock_count": len(stock_codes),
                "period_results": period_results,
                "common_features": common_features,
                "high_frequency_features": high_frequency_features,
                "pattern_features": pattern_features,
                "trend_features": trend_features,
                "volume_features": volume_features,
                "feature_correlation": correlation_result,
                "temporal_patterns": temporal_result,
                "optimized_features": optimized_features
            }
            
            logger.info(f"多周期分析完成，发现 {len(common_features)} 个共性特征，{len(correlation_result.get('strong_associations', []))} 个强关联规则")
            return result
            
        except Exception as e:
            logger.error(f"多周期分析时出错: {e}")
            return {}
    
    def analyze_indicator_signals(self, stock_codes: List[str], 
                                start_date: str, end_date: str,
                                indicators: List[Dict[str, Any]],
                                period: str = 'DAILY') -> Dict[str, Any]:
        """
        分析指标信号
        
        Args:
            stock_codes: 股票代码列表
            start_date: 开始日期，格式YYYYMMDD
            end_date: 结束日期，格式YYYYMMDD
            indicators: 指标列表，每个指标为Dict格式
                {
                    "name": "指标名称",
                    "id": "指标ID",
                    "parameters": {"参数1": 值1, ...}
                }
            period: K线周期
            
        Returns:
            Dict: 指标信号分析结果
        """
        logger.info(f"开始分析指标信号: {len(stock_codes)}只股票, {len(indicators)}个指标, 周期={period}")
        
        try:
            # 指标信号统计
            indicator_stats = {}
            
            # 对每个股票计算指标
            for stock_code in stock_codes:
                # 获取股票K线数据
                start_date_extended = self._get_extended_date(start_date, days=50)  # 预留足够数据计算指标
                
                sql = f"""
                SELECT date, open, high, low, close, volume, amount
                FROM stock_{period.lower()}
                WHERE stock_code = '{stock_code}'
                  AND date >= '{start_date_extended}' AND date <= '{end_date}'
                ORDER BY date
                """
                kline_data = self.ch_db.query_df(sql)
                
                if kline_data.empty:
                    logger.warning(f"未找到股票 {stock_code} 的K线数据")
                    continue
                
                # 过滤掉扩展的日期数据
                valid_data = kline_data[kline_data['date'] >= start_date]
                
                if valid_data.empty:
                    continue
                
                # 对每个指标计算信号
                for indicator_config in indicators:
                    indicator_id = indicator_config["id"]
                    indicator_name = indicator_config["name"]
                    parameters = indicator_config.get("parameters", {})
                    
                    # 统计字典初始化
                    if indicator_id not in indicator_stats:
                        indicator_stats[indicator_id] = {
                            "name": indicator_name,
                            "signals": {}
                        }
                    
                    try:
                        # 创建指标实例
                        indicator = self.indicator_factory.create(indicator_id, **parameters)
                        
                        # 计算指标值
                        indicator_result = indicator.calculate(kline_data)
                        
                        # 生成信号
                        signals = indicator.generate_signals(indicator_result)
                        
                        # 过滤有效期内的信号
                        valid_signals = signals[signals.index.isin(valid_data.index)]
                        
                        # 统计信号出现次数
                        for column in valid_signals.columns:
                            # 只统计布尔型信号列
                            if valid_signals[column].dtype == bool:
                                signal_count = valid_signals[column].sum()
                                
                                if column not in indicator_stats[indicator_id]["signals"]:
                                    indicator_stats[indicator_id]["signals"][column] = 0
                                
                                indicator_stats[indicator_id]["signals"][column] += signal_count
                                
                    except Exception as e:
                        logger.error(f"计算指标 {indicator_id} 时出错: {e}")
            
            # 计算信号频率
            result = {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "stock_count": len(stock_codes),
                "indicator_stats": self._calculate_signal_frequency(indicator_stats, len(stock_codes))
            }
            
            # 识别高频信号
            result["common_signals"] = self._identify_common_signals(result["indicator_stats"])
            
            # 保存到分析结果
            self.analysis_results["indicator_analysis"] = result
            
            logger.info(f"指标信号分析完成")
            return result
            
        except Exception as e:
            logger.error(f"分析指标信号时出错: {e}")
            return {}
    
    def _get_period_date_range(self, date: str, period: str) -> Tuple[str, str]:
        """根据周期获取日期范围"""
        date_obj = datetime.strptime(date, "%Y%m%d")
        
        if period == 'DAILY':
            # 日线取最近30天
            start_date = (date_obj - timedelta(days=30)).strftime("%Y%m%d")
            end_date = date
        elif period == 'WEEKLY':
            # 周线取最近12周
            start_date = (date_obj - timedelta(days=12*7)).strftime("%Y%m%d")
            end_date = date
        elif period == 'MONTHLY':
            # 月线取最近6个月
            start_date = (date_obj - timedelta(days=6*30)).strftime("%Y%m%d")
            end_date = date
        else:
            # 默认取30天
            start_date = (date_obj - timedelta(days=30)).strftime("%Y%m%d")
            end_date = date
            
        return start_date, end_date
    
    def _find_multi_period_common_features(self, period_results: Dict[str, Dict[str, List[Dict[str, Any]]]]) -> List[Dict[str, Any]]:
        """寻找多周期共性特征"""
        # 将所有特征平铺并记录周期信息
        all_features = []
        
        for period, features in period_results.items():
            for feature_type, feature_list in features.items():
                for feature in feature_list:
                    all_features.append({
                        "period": period,
                        "type": feature_type,
                        "feature": feature["pattern"] if "pattern" in feature else feature.get("signal", ""),
                        "frequency": feature["frequency"],
                        "category": feature.get("category", ""),
                        "count": feature.get("count", 0)
                    })
        
        # 按特征名称分组
        feature_groups = {}
        for feature in all_features:
            key = f"{feature['type']}_{feature['feature']}"
            if key not in feature_groups:
                feature_groups[key] = []
            feature_groups[key].append(feature)
        
        # 寻找出现在多个周期的特征
        common_features = []
        for key, group in feature_groups.items():
            if len(group) > 1:  # 出现在多个周期
                periods = sorted(list(set([f["period"] for f in group])))
                avg_frequency = sum([f["frequency"] for f in group]) / len(group)
                
                common_features.append({
                    "feature": group[0]["feature"],
                    "type": group[0]["type"],
                    "category": group[0]["category"],
                    "periods": periods,
                    "avg_frequency": avg_frequency,
                    "details": group
                })
        
        # 按平均频率排序
        common_features.sort(key=lambda x: x["avg_frequency"], reverse=True)
        
        return common_features
    
    def _calculate_signal_frequency(self, indicator_stats: Dict[str, Dict[str, Any]], 
                                  stock_count: int) -> Dict[str, Any]:
        """计算信号频率"""
        result = {}
        
        for indicator_id, stats in indicator_stats.items():
            signals_result = {}
            
            for signal, count in stats["signals"].items():
                frequency = count / stock_count if stock_count > 0 else 0
                signals_result[signal] = {
                    "count": count,
                    "frequency": frequency
                }
            
            # 按频率排序
            sorted_signals = {k: v for k, v in sorted(
                signals_result.items(), 
                key=lambda item: item[1]["frequency"], 
                reverse=True
            )}
            
            result[indicator_id] = {
                "name": stats["name"],
                "signals": sorted_signals
            }
        
        return result
    
    def _identify_common_signals(self, indicator_stats: Dict[str, Any]) -> List[Dict[str, Any]]:
        """识别共性信号特征"""
        common_signals = []
        
        # 设置频率阈值
        threshold = 0.3
        
        # 遍历各指标信号
        for indicator_id, stats in indicator_stats.items():
            for signal, signal_stats in stats["signals"].items():
                if signal_stats["frequency"] >= threshold:
                    common_signals.append({
                        "indicator": stats["name"],
                        "indicator_id": indicator_id,
                        "signal": signal,
                        "frequency": signal_stats["frequency"],
                        "count": signal_stats["count"]
                    })
        
        # 按频率排序
        common_signals.sort(key=lambda x: x["frequency"], reverse=True)
        
        return common_signals

    def get_comprehensive_report(self) -> Dict[str, Any]:
        """
        获取买点分析综合报告
        
        Returns:
            Dict: 综合分析报告
        """
        logger.info("生成买点维度综合分析报告")
        
        try:
            # 整合各模块分析结果
            report = {
                "report_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "pattern_analysis": {
                    "common_features": self.analysis_results.get("pattern_analysis", {}).get("common_features", [])
                },
                "trend_analysis": {
                    "common_features": self.analysis_results.get("trend_analysis", {}).get("common_features", [])
                },
                "volume_analysis": {
                    "common_features": self.analysis_results.get("volume_analysis", {}).get("common_features", [])
                },
                "indicator_analysis": {
                    "common_signals": self.analysis_results.get("indicator_analysis", {}).get("common_signals", [])
                },
                "multi_period_analysis": {
                    "common_features": self.analysis_results.get("multi_period_analysis", {}).get("common_features", []),
                    "feature_correlation": self.analysis_results.get("multi_period_analysis", {}).get("feature_correlation", {}),
                    "temporal_patterns": self.analysis_results.get("multi_period_analysis", {}).get("temporal_patterns", {}),
                    "optimized_features": self.analysis_results.get("multi_period_analysis", {}).get("optimized_features", {})
                }
            }
            
            # 构建买点关键特征
            key_features = []
            
            # 从各分析中提取高频特征
            for analysis_type in ["pattern_analysis", "trend_analysis", "volume_analysis"]:
                features = self.analysis_results.get(analysis_type, {}).get("common_features", [])
                for feature in features:
                    if isinstance(feature, dict) and "frequency" in feature and feature["frequency"] > 0.5:
                        key_features.append({
                            "type": analysis_type.replace("_analysis", ""),
                            "feature": feature.get("pattern", feature.get("signal", "")),
                            "frequency": feature["frequency"],
                            "count": feature.get("count", 0)
                        })
            
            # 添加指标信号
            signals = self.analysis_results.get("indicator_analysis", {}).get("common_signals", [])
            for signal in signals:
                if signal["frequency"] > 0.5:
                    key_features.append({
                        "type": "indicator",
                        "feature": f"{signal['indicator']}:{signal['signal']}",
                        "frequency": signal["frequency"],
                        "count": signal.get("count", 0)
                    })
            
            # 添加优化特征组合
            optimized_features = self.analysis_results.get("multi_period_analysis", {}).get("optimized_features", {}).get("optimized_features", [])
            if optimized_features:
                for feature in optimized_features:
                    if "frequency" in feature and feature["frequency"] > 0.3:
                        key_features.append({
                            "type": feature.get("type", "optimized"),
                            "feature": feature.get("pattern", feature.get("signal", feature.get("feature", ""))),
                            "frequency": feature["frequency"],
                            "count": feature.get("count", 0),
                            "source": "optimized"
                        })
            
            # 添加强关联规则
            strong_associations = self.analysis_results.get("multi_period_analysis", {}).get("feature_correlation", {}).get("strong_associations", [])
            if strong_associations:
                for i, rule in enumerate(strong_associations[:5]):  # 只添加前5个强关联规则
                    key_features.append({
                        "type": "association_rule",
                        "feature": f"{rule['antecedent']} -> {rule['consequent']}",
                        "frequency": rule.get("support", 0),
                        "confidence": rule.get("confidence", 0),
                        "lift": rule.get("lift", 0),
                        "source": "correlation"
                    })
            
            # 添加时序模式
            frequent_patterns = self.analysis_results.get("multi_period_analysis", {}).get("temporal_patterns", {}).get("frequent_patterns", [])
            if frequent_patterns:
                for i, pattern in enumerate(frequent_patterns[:3]):  # 只添加前3个频繁时序模式
                    sequence = pattern.get("sequence", [])
                    if sequence:
                        sequence_str = " -> ".join([f"{item.get('type', '')}:{item.get('feature', '')}" for item in sequence])
                        key_features.append({
                            "type": "temporal_pattern",
                            "feature": sequence_str,
                            "support": pattern.get("support", 0),
                            "count": pattern.get("count", 0),
                            "source": "temporal"
                        })
            
            # 按频率排序
            key_features.sort(key=lambda x: x.get("frequency", x.get("support", 0)), reverse=True)
            
            report["key_features"] = key_features
            
            # 买点描述生成
            description = self._generate_buypoint_description(key_features)
            report["description"] = description
            
            # 特征评分
            report["feature_scores"] = self._score_features(key_features)
            
            # 特征关联网络
            report["feature_network"] = self.analysis_results.get("multi_period_analysis", {}).get("feature_correlation", {}).get("feature_network", [])
            
            # 优化特征组合
            feature_combinations = self.analysis_results.get("multi_period_analysis", {}).get("optimized_features", {}).get("feature_combinations", [])
            report["feature_combinations"] = feature_combinations
            
            logger.info("买点维度综合分析报告生成完成")
            return report
            
        except Exception as e:
            logger.error(f"生成买点维度综合分析报告时出错: {e}")
            return {}
    
    def _score_features(self, features: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        对特征进行评分
        
        Args:
            features: 特征列表
            
        Returns:
            Dict: 特征评分结果
        """
        try:
            # 特征类型权重
            type_weights = {
                "pattern": 1.0,
                "trend": 1.2,
                "volume": 0.9,
                "indicator": 1.1,
                "association_rule": 1.3,
                "temporal_pattern": 1.4,
                "optimized": 1.5
            }
            
            # 计算各维度得分
            dimension_scores = {
                "pattern": 0,
                "trend": 0,
                "volume": 0,
                "indicator": 0,
                "correlation": 0,
                "temporal": 0
            }
            
            dimension_counts = {dim: 0 for dim in dimension_scores.keys()}
            
            # 计算每个特征的加权得分
            feature_scores = {}
            for feature in features:
                feature_type = feature.get("type", "")
                feature_name = feature.get("feature", "")
                
                # 特征基础分数
                base_score = feature.get("frequency", feature.get("support", 0)) * 100
                
                # 应用类型权重
                type_weight = type_weights.get(feature_type, 1.0)
                weighted_score = base_score * type_weight
                
                # 特征来源加成
                source = feature.get("source", "")
                if source == "optimized":
                    weighted_score *= 1.2
                elif source == "correlation":
                    weighted_score *= 1.15
                elif source == "temporal":
                    weighted_score *= 1.25
                
                # 存储特征得分
                feature_scores[feature_name] = weighted_score
                
                # 更新维度得分
                if feature_type in ["pattern", "trend", "volume", "indicator"]:
                    dimension_scores[feature_type] += weighted_score
                    dimension_counts[feature_type] += 1
                elif feature_type == "association_rule":
                    dimension_scores["correlation"] += weighted_score
                    dimension_counts["correlation"] += 1
                elif feature_type == "temporal_pattern":
                    dimension_scores["temporal"] += weighted_score
                    dimension_counts["temporal"] += 1
            
            # 计算维度平均分
            for dim in dimension_scores:
                if dimension_counts[dim] > 0:
                    dimension_scores[dim] /= dimension_counts[dim]
                    
            # 计算总得分
            total_score = sum(dimension_scores.values()) / len([s for s in dimension_scores.values() if s > 0]) if any(dimension_scores.values()) else 0
            
            # 结果
            result = {
                "total_score": total_score,
                "dimension_scores": dimension_scores,
                "feature_scores": feature_scores
            }
            
            return result
            
        except Exception as e:
            logger.error(f"特征评分时出错: {e}")
            return {
                "total_score": 0,
                "dimension_scores": {},
                "feature_scores": {}
            }
    
    def save_results(self, output_file: str, format_type: str = "json") -> None:
        """
        保存分析结果
        
        Args:
            output_file: 输出文件路径
            format_type: 输出格式类型，支持json和markdown
        """
        logger.info(f"保存买点分析结果到: {output_file}")
        
        try:
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            if format_type.lower() == "json":
                # 获取综合报告
                report = self.get_comprehensive_report()
                
                # 构建完整结果
                full_results = {
                    "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "comprehensive_report": report,
                    "detailed_results": self.analysis_results
                }
                
                # 保存到文件
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(full_results, f, indent=2, ensure_ascii=False)
                
            elif format_type.lower() == "markdown":
                # 获取综合报告
                report = self.get_comprehensive_report()
                
                # 构建Markdown内容
                md_content = self._generate_markdown_report(report)
                
                # 保存到文件
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(md_content)
            
            logger.info(f"买点分析结果已保存到: {output_file}")
            
        except Exception as e:
            logger.error(f"保存买点分析结果时出错: {e}")
    
    def _generate_markdown_report(self, report: Dict[str, Any]) -> str:
        """
        生成Markdown格式的报告
        
        Args:
            report: 报告数据
            
        Returns:
            str: Markdown格式的报告内容
        """
        md = []
        
        # 添加标题
        md.append("# 买点维度分析报告\n")
        md.append(f"*报告生成时间: {report.get('report_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}*\n")
        
        # 添加买点描述
        md.append("## 买点描述\n")
        md.append(report.get("description", "无法生成买点描述，关键特征不足。"))
        md.append("\n")
        
        # 添加特征评分
        feature_scores = report.get("feature_scores", {})
        if feature_scores:
            md.append("## 特征维度评分\n")
            md.append(f"**总评分**: {feature_scores.get('total_score', 0):.2f}\n")
            
            md.append("### 各维度评分\n")
            dimension_scores = feature_scores.get("dimension_scores", {})
            md.append("| 维度 | 评分 |\n")
            md.append("| --- | --- |\n")
            for dim, score in dimension_scores.items():
                md.append(f"| {dim} | {score:.2f} |\n")
            md.append("\n")
        
        # 添加关键特征
        md.append("## 关键特征\n")
        key_features = report.get("key_features", [])
        if key_features:
            md.append("| 特征类型 | 特征内容 | 频率/支持度 | 计数 |\n")
            md.append("| --- | --- | --- | --- |\n")
            
            for feature in key_features:
                feature_type = feature.get("type", "")
                feature_name = feature.get("feature", "")
                frequency = feature.get("frequency", feature.get("support", 0))
                count = feature.get("count", 0)
                
                md.append(f"| {feature_type} | {feature_name} | {frequency:.2f} | {count} |\n")
            md.append("\n")
        
        # 添加特征关联分析
        md.append("## 特征关联分析\n")
        feature_correlation = report.get("multi_period_analysis", {}).get("feature_correlation", {})
        if feature_correlation:
            # 强关联规则
            strong_associations = feature_correlation.get("strong_associations", [])
            if strong_associations:
                md.append("### 强关联规则\n")
                md.append("| 前项 | 后项 | 支持度 | 置信度 | 提升度 |\n")
                md.append("| --- | --- | --- | --- | --- |\n")
                
                for rule in strong_associations[:10]:  # 只显示前10个规则
                    antecedent = rule.get("antecedent", "")
                    consequent = rule.get("consequent", "")
                    support = rule.get("support", 0)
                    confidence = rule.get("confidence", 0)
                    lift = rule.get("lift", 0)
                    
                    md.append(f"| {antecedent} | {consequent} | {support:.2f} | {confidence:.2f} | {lift:.2f} |\n")
                md.append("\n")
            
            # 特征重要性
            feature_importance = feature_correlation.get("feature_importance", {})
            if feature_importance:
                md.append("### 特征重要性\n")
                md.append("| 特征 | 重要性分数 |\n")
                md.append("| --- | --- |\n")
                
                for feature, score in list(feature_importance.items())[:15]:  # 只显示前15个特征
                    md.append(f"| {feature} | {score:.2f} |\n")
                md.append("\n")
        
        # 添加时序模式分析
        md.append("## 时序模式分析\n")
        temporal_patterns = report.get("multi_period_analysis", {}).get("temporal_patterns", {})
        if temporal_patterns:
            frequent_patterns = temporal_patterns.get("frequent_patterns", [])
            if frequent_patterns:
                md.append("### 频繁时序模式\n")
                md.append("| 时序模式 | 支持度 | 计数 |\n")
                md.append("| --- | --- | --- |\n")
                
                for pattern in frequent_patterns[:10]:  # 只显示前10个模式
                    sequence = pattern.get("sequence", [])
                    if sequence:
                        sequence_str = " -> ".join([f"{item.get('type', '')}:{item.get('feature', '')}" for item in sequence])
                        support = pattern.get("support", 0)
                        count = pattern.get("count", 0)
                        
                        md.append(f"| {sequence_str} | {support:.2f} | {count} |\n")
                md.append("\n")
        
        # 添加优化特征组合
        md.append("## 优化特征组合\n")
        optimized_features = report.get("multi_period_analysis", {}).get("optimized_features", {})
        if optimized_features:
            feature_list = optimized_features.get("optimized_features", [])
            if feature_list:
                md.append("### 最优特征集\n")
                md.append("| 特征 | 频率 | 来源 |\n")
                md.append("| --- | --- | --- |\n")
                
                for feature in feature_list:
                    feature_name = feature.get("pattern", feature.get("signal", feature.get("feature", "")))
                    frequency = feature.get("frequency", 0)
                    feature_type = feature.get("type", "")
                    
                    md.append(f"| {feature_name} | {frequency:.2f} | {feature_type} |\n")
                md.append("\n")
            
            # 特征组合
            feature_combinations = report.get("feature_combinations", [])
            if feature_combinations:
                md.append("### 推荐特征组合\n")
                md.append("| 组合描述 | 组合类型 |\n")
                md.append("| --- | --- |\n")
                
                for combo in feature_combinations[:7]:  # 只显示前7个组合
                    description = combo.get("description", "")
                    combo_type = combo.get("combo_type", "")
                    
                    md.append(f"| {description} | {combo_type} |\n")
                md.append("\n")
        
        # 添加各周期分析结果摘要
        md.append("## 各周期分析摘要\n")
        period_results = report.get("multi_period_analysis", {}).get("period_results", {})
        if period_results:
            for period, results in period_results.items():
                md.append(f"### {period} 周期\n")
                
                # 形态特征
                pattern_features = results.get("pattern_features", [])
                if pattern_features:
                    md.append("#### 形态特征\n")
                    md.append("| 特征 | 频率 | 计数 |\n")
                    md.append("| --- | --- | --- |\n")
                    
                    for feature in pattern_features[:5]:  # 只显示前5个特征
                        pattern = feature.get("pattern", "")
                        frequency = feature.get("frequency", 0)
                        count = feature.get("count", 0)
                        
                        md.append(f"| {pattern} | {frequency:.2f} | {count} |\n")
                    md.append("\n")
                
                # 趋势特征
                trend_features = results.get("trend_features", [])
                if trend_features:
                    md.append("#### 趋势特征\n")
                    md.append("| 特征 | 频率 | 计数 |\n")
                    md.append("| --- | --- | --- |\n")
                    
                    for feature in trend_features[:5]:  # 只显示前5个特征
                        pattern = feature.get("pattern", "")
                        frequency = feature.get("frequency", 0)
                        count = feature.get("count", 0)
                        
                        md.append(f"| {pattern} | {frequency:.2f} | {count} |\n")
                    md.append("\n")
                
                # 量能特征
                volume_features = results.get("volume_features", [])
                if volume_features:
                    md.append("#### 量能特征\n")
                    md.append("| 特征 | 频率 | 计数 |\n")
                    md.append("| --- | --- | --- |\n")
                    
                    for feature in volume_features[:5]:  # 只显示前5个特征
                        pattern = feature.get("pattern", "")
                        frequency = feature.get("frequency", 0)
                        count = feature.get("count", 0)
                        
                        md.append(f"| {pattern} | {frequency:.2f} | {count} |\n")
                    md.append("\n")
        
        return "".join(md)

    def export_to_excel(self, output_file: str) -> None:
        """
        导出分析结果到Excel文件
        
        Args:
            output_file: 输出文件路径
        """
        logger.info(f"导出买点分析结果到Excel: {output_file}")
        
        try:
            # 获取综合报告
            report = self.get_comprehensive_report()
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 创建概览工作表
            overview_sheet = wb.active
            overview_sheet.title = "分析概览"
            
            # 添加标题
            overview_sheet['A1'] = "买点维度分析报告"
            overview_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            overview_sheet.merge_cells('A1:E1')
            
            # 添加报告时间
            overview_sheet['A2'] = "报告生成时间:"
            overview_sheet['B2'] = report.get('report_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            # 添加买点描述
            overview_sheet['A4'] = "买点描述"
            overview_sheet['A4'].font = openpyxl.styles.Font(bold=True)
            overview_sheet.merge_cells('A4:E4')
            
            overview_sheet['A5'] = report.get("description", "无法生成买点描述，关键特征不足。")
            overview_sheet.merge_cells('A5:E5')
            overview_sheet.row_dimensions[5].height = 60
            
            # 添加特征评分
            overview_sheet['A7'] = "特征评分"
            overview_sheet['A7'].font = openpyxl.styles.Font(bold=True)
            overview_sheet.merge_cells('A7:E7')
            
            feature_scores = report.get("feature_scores", {})
            if feature_scores:
                overview_sheet['A8'] = "总评分:"
                overview_sheet['B8'] = feature_scores.get('total_score', 0)
                
                # 各维度评分
                overview_sheet['A10'] = "各维度评分:"
                overview_sheet['A10'].font = openpyxl.styles.Font(bold=True)
                overview_sheet.merge_cells('A10:E10')
                
                overview_sheet['A11'] = "维度"
                overview_sheet['B11'] = "评分"
                overview_sheet['A11'].font = openpyxl.styles.Font(bold=True)
                overview_sheet['B11'].font = openpyxl.styles.Font(bold=True)
                
                dimension_scores = feature_scores.get("dimension_scores", {})
                row = 12
                for dim, score in dimension_scores.items():
                    overview_sheet[f'A{row}'] = dim
                    overview_sheet[f'B{row}'] = score
                    row += 1
            
            # 创建关键特征工作表
            key_features_sheet = wb.create_sheet("关键特征")
            
            key_features_sheet['A1'] = "关键特征"
            key_features_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            key_features_sheet.merge_cells('A1:E1')
            
            key_features_sheet['A3'] = "特征类型"
            key_features_sheet['B3'] = "特征内容"
            key_features_sheet['C3'] = "频率/支持度"
            key_features_sheet['D3'] = "计数"
            
            for col in ['A', 'B', 'C', 'D']:
                key_features_sheet[f'{col}3'].font = openpyxl.styles.Font(bold=True)
            
            key_features = report.get("key_features", [])
            row = 4
            for feature in key_features:
                key_features_sheet[f'A{row}'] = feature.get("type", "")
                key_features_sheet[f'B{row}'] = feature.get("feature", "")
                key_features_sheet[f'C{row}'] = feature.get("frequency", feature.get("support", 0))
                key_features_sheet[f'D{row}'] = feature.get("count", 0)
                row += 1
            
            # 创建特征关联分析工作表
            correlation_sheet = wb.create_sheet("特征关联分析")
            
            correlation_sheet['A1'] = "特征关联分析"
            correlation_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            correlation_sheet.merge_cells('A1:E1')
            
            feature_correlation = report.get("multi_period_analysis", {}).get("feature_correlation", {})
            if feature_correlation:
                # 强关联规则
                correlation_sheet['A3'] = "强关联规则"
                correlation_sheet['A3'].font = openpyxl.styles.Font(bold=True)
                correlation_sheet.merge_cells('A3:E3')
                
                correlation_sheet['A4'] = "前项"
                correlation_sheet['B4'] = "后项"
                correlation_sheet['C4'] = "支持度"
                correlation_sheet['D4'] = "置信度"
                correlation_sheet['E4'] = "提升度"
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    correlation_sheet[f'{col}4'].font = openpyxl.styles.Font(bold=True)
                
                strong_associations = feature_correlation.get("strong_associations", [])
                row = 5
                for rule in strong_associations:
                    correlation_sheet[f'A{row}'] = rule.get("antecedent", "")
                    correlation_sheet[f'B{row}'] = rule.get("consequent", "")
                    correlation_sheet[f'C{row}'] = rule.get("support", 0)
                    correlation_sheet[f'D{row}'] = rule.get("confidence", 0)
                    correlation_sheet[f'E{row}'] = rule.get("lift", 0)
                    row += 1
                
                # 特征重要性
                row += 2
                correlation_sheet[f'A{row}'] = "特征重要性"
                correlation_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                correlation_sheet.merge_cells(f'A{row}:C{row}')
                
                row += 1
                correlation_sheet[f'A{row}'] = "特征"
                correlation_sheet[f'B{row}'] = "重要性分数"
                correlation_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                correlation_sheet[f'B{row}'].font = openpyxl.styles.Font(bold=True)
                
                feature_importance = feature_correlation.get("feature_importance", {})
                row += 1
                for feature, score in list(feature_importance.items())[:20]:  # 只显示前20个特征
                    correlation_sheet[f'A{row}'] = feature
                    correlation_sheet[f'B{row}'] = score
                    row += 1
            
            # 创建时序模式分析工作表
            temporal_sheet = wb.create_sheet("时序模式分析")
            
            temporal_sheet['A1'] = "时序模式分析"
            temporal_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            temporal_sheet.merge_cells('A1:C1')
            
            temporal_patterns = report.get("multi_period_analysis", {}).get("temporal_patterns", {})
            if temporal_patterns:
                frequent_patterns = temporal_patterns.get("frequent_patterns", [])
                if frequent_patterns:
                    temporal_sheet['A3'] = "频繁时序模式"
                    temporal_sheet['A3'].font = openpyxl.styles.Font(bold=True)
                    temporal_sheet.merge_cells('A3:C3')
                    
                    temporal_sheet['A4'] = "时序模式"
                    temporal_sheet['B4'] = "支持度"
                    temporal_sheet['C4'] = "计数"
                    
                    for col in ['A', 'B', 'C']:
                        temporal_sheet[f'{col}4'].font = openpyxl.styles.Font(bold=True)
                    
                    row = 5
                    for pattern in frequent_patterns:
                        sequence = pattern.get("sequence", [])
                        if sequence:
                            sequence_str = " -> ".join([f"{item.get('type', '')}:{item.get('feature', '')}" for item in sequence])
                            temporal_sheet[f'A{row}'] = sequence_str
                            temporal_sheet[f'B{row}'] = pattern.get("support", 0)
                            temporal_sheet[f'C{row}'] = pattern.get("count", 0)
                            row += 1
            
            # 创建优化特征组合工作表
            optimized_sheet = wb.create_sheet("优化特征组合")
            
            optimized_sheet['A1'] = "优化特征组合"
            optimized_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            optimized_sheet.merge_cells('A1:C1')
            
            optimized_features = report.get("multi_period_analysis", {}).get("optimized_features", {})
            if optimized_features:
                feature_list = optimized_features.get("optimized_features", [])
                if feature_list:
                    optimized_sheet['A3'] = "最优特征集"
                    optimized_sheet['A3'].font = openpyxl.styles.Font(bold=True)
                    optimized_sheet.merge_cells('A3:C3')
                    
                    optimized_sheet['A4'] = "特征"
                    optimized_sheet['B4'] = "频率"
                    optimized_sheet['C4'] = "类型"
                    
                    for col in ['A', 'B', 'C']:
                        optimized_sheet[f'{col}4'].font = openpyxl.styles.Font(bold=True)
                    
                    row = 5
                    for feature in feature_list:
                        feature_name = feature.get("pattern", feature.get("signal", feature.get("feature", "")))
                        optimized_sheet[f'A{row}'] = feature_name
                        optimized_sheet[f'B{row}'] = feature.get("frequency", 0)
                        optimized_sheet[f'C{row}'] = feature.get("type", "")
                        row += 1
                
                # 特征组合
                feature_combinations = optimized_features.get("feature_combinations", [])
                if feature_combinations:
                    row += 2
                    optimized_sheet[f'A{row}'] = "推荐特征组合"
                    optimized_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                    optimized_sheet.merge_cells(f'A{row}:C{row}')
                    
                    row += 1
                    optimized_sheet[f'A{row}'] = "组合描述"
                    optimized_sheet[f'B{row}'] = "组合类型"
                    optimized_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                    optimized_sheet[f'B{row}'].font = openpyxl.styles.Font(bold=True)
                    
                    row += 1
                    for combo in feature_combinations:
                        optimized_sheet[f'A{row}'] = combo.get("description", "")
                        optimized_sheet[f'B{row}'] = combo.get("combo_type", "")
                        row += 1
            
            # 调整列宽
            for sheet in wb.worksheets:
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column].width = min(adjusted_width, 60)
            
            # 保存Excel文件
            wb.save(output_file)
            logger.info(f"买点分析结果已导出到Excel: {output_file}")
            
        except Exception as e:
            logger.error(f"导出买点分析结果到Excel时出错: {e}")
    
    def _generate_buypoint_description(self, key_features: List[Dict[str, Any]]) -> str:
        """
        生成买点描述
        
        Args:
            key_features: 关键特征列表
            
        Returns:
            str: 买点描述文本
        """
        if not key_features:
            return "无法生成买点描述，关键特征不足。"
        
        # 按类型分组特征
        feature_by_type = {}
        for feature in key_features:
            feature_type = feature["type"]
            if feature_type not in feature_by_type:
                feature_by_type[feature_type] = []
            feature_by_type[feature_type].append(feature)
        
        description_parts = []
        
        # 形态描述
        if "pattern" in feature_by_type:
            pattern_features = feature_by_type["pattern"]
            if pattern_features:
                patterns = [f"{f['feature']}({f['frequency']*100:.0f}%)" for f in pattern_features[:3]]
                description_parts.append(f"K线形态呈现{', '.join(patterns)}")
        
        # 趋势描述
        if "trend" in feature_by_type:
            trend_features = feature_by_type["trend"]
            if trend_features:
                trends = [f"{f['feature']}({f['frequency']*100:.0f}%)" for f in trend_features[:3]]
                description_parts.append(f"趋势特征表现为{', '.join(trends)}")
        
        # 量能描述
        if "volume" in feature_by_type:
            volume_features = feature_by_type["volume"]
            if volume_features:
                volumes = [f"{f['feature']}({f['frequency']*100:.0f}%)" for f in volume_features[:3]]
                description_parts.append(f"量能表现为{', '.join(volumes)}")
        
        # 指标信号描述
        if "indicator" in feature_by_type:
            indicator_features = feature_by_type["indicator"]
            if indicator_features:
                indicators = [f"{f['feature']}({f['frequency']*100:.0f}%)" for f in indicator_features[:3]]
                description_parts.append(f"技术指标显示{', '.join(indicators)}")
        
        # 综合描述
        if description_parts:
            description = "该买点特征为：" + "；".join(description_parts) + "。"
            
            # 添加总结
            top_features = sorted(key_features, key=lambda x: x["frequency"], reverse=True)[:5]
            if top_features:
                top_feature_texts = [f"{f['feature']}({f['frequency']*100:.0f}%)" for f in top_features]
                description += f"\n\n综合来看，最显著的特征是：{', '.join(top_feature_texts)}。"
            
            return description
        else:
            return "无法生成买点描述，关键特征不足。" 

    def analyze_feature_correlation(self, features_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        分析特征之间的关联性
        
        Args:
            features_data: 特征数据列表，每个特征包含pattern/signal和出现情况
            
        Returns:
            Dict: 特征关联分析结果
        """
        logger.info("开始分析特征之间的关联性")
        
        try:
            # 按股票和日期组织特征出现情况
            feature_occurrences = {}
            
            for feature in features_data:
                stock_code = feature.get("stock_code", "")
                date = feature.get("date", "")
                feature_name = feature.get("pattern", feature.get("signal", ""))
                feature_type = feature.get("type", "")
                
                key = f"{stock_code}_{date}"
                if key not in feature_occurrences:
                    feature_occurrences[key] = set()
                
                feature_occurrences[key].add(f"{feature_type}:{feature_name}")
            
            # 构建特征共现矩阵
            all_features = set()
            for features in feature_occurrences.values():
                all_features.update(features)
            
            all_features = sorted(list(all_features))
            feature_index = {feature: i for i, feature in enumerate(all_features)}
            
            # 初始化共现矩阵
            co_occurrence_matrix = np.zeros((len(all_features), len(all_features)), dtype=int)
            
            # 填充共现矩阵
            for features in feature_occurrences.values():
                for feature1 in features:
                    for feature2 in features:
                        if feature1 != feature2:
                            i = feature_index[feature1]
                            j = feature_index[feature2]
                            co_occurrence_matrix[i, j] += 1
            
            # 计算关联规则
            association_rules = []
            total_records = len(feature_occurrences)
            
            for i, feature1 in enumerate(all_features):
                for j, feature2 in enumerate(all_features):
                    if i != j and co_occurrence_matrix[i, j] > 0:
                        # 计算支持度和置信度
                        support = co_occurrence_matrix[i, j] / total_records
                        
                        # 计算feature1的出现次数
                        feature1_count = sum(1 for features in feature_occurrences.values() if feature1 in features)
                        
                        confidence = co_occurrence_matrix[i, j] / feature1_count if feature1_count > 0 else 0
                        
                        # 计算提升度
                        feature2_count = sum(1 for features in feature_occurrences.values() if feature2 in features)
                        expected_confidence = feature2_count / total_records if total_records > 0 else 0
                        lift = confidence / expected_confidence if expected_confidence > 0 else 0
                        
                        association_rules.append({
                            "antecedent": feature1,
                            "consequent": feature2,
                            "support": support,
                            "confidence": confidence,
                            "lift": lift,
                            "co_occurrence_count": int(co_occurrence_matrix[i, j])
                        })
            
            # 按提升度排序
            association_rules.sort(key=lambda x: x["lift"], reverse=True)
            
            # 提取强关联规则
            strong_rules = [rule for rule in association_rules if rule["lift"] > 1.5 and rule["confidence"] > 0.5]
            
            # 构建特征关联网络
            feature_network = []
            for rule in association_rules[:min(100, len(association_rules))]:
                feature_network.append({
                    "source": rule["antecedent"],
                    "target": rule["consequent"],
                    "weight": rule["lift"]
                })
            
            # 计算特征重要性（基于PageRank算法思想）
            feature_importance = self._calculate_feature_importance(all_features, association_rules)
            
            # 保存分析结果
            correlation_result = {
                "feature_count": len(all_features),
                "record_count": total_records,
                "strong_associations": strong_rules[:20],  # 只保留前20个强关联规则
                "feature_network": feature_network,
                "feature_importance": feature_importance
            }
            
            logger.info(f"特征关联性分析完成，发现 {len(strong_rules)} 个强关联规则")
            return correlation_result
            
        except Exception as e:
            logger.error(f"分析特征关联性时出错: {e}")
            return {}
    
    def _calculate_feature_importance(self, features: List[str], 
                                   association_rules: List[Dict[str, Any]]) -> Dict[str, float]:
        """计算特征重要性（基于PageRank算法思想）"""
        # 初始化每个特征的重要性为1
        importance = {feature: 1.0 for feature in features}
        
        # 迭代计算
        damping = 0.85
        iterations = 10
        
        for _ in range(iterations):
            new_importance = {feature: 0.0 for feature in features}
            
            # 计算特征转移矩阵
            for rule in association_rules:
                source = rule["antecedent"]
                target = rule["consequent"]
                weight = rule["lift"]
                
                # 累加来源特征的加权重要性
                new_importance[target] += importance[source] * weight
            
            # 应用阻尼因子并归一化
            total = sum(new_importance.values())
            if total > 0:
                for feature in features:
                    new_importance[feature] = (1 - damping) + damping * new_importance[feature] / total
            
            importance = new_importance
        
        # 排序并返回
        sorted_importance = {k: v for k, v in sorted(
            importance.items(), key=lambda item: item[1], reverse=True
        )}
        
        return sorted_importance
    
    def analyze_temporal_patterns(self, features_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        分析特征的时序模式
        
        Args:
            features_data: 特征数据列表，每个特征包含pattern/signal和出现日期
            
        Returns:
            Dict: 时序模式分析结果
        """
        logger.info("开始分析特征的时序模式")
        
        try:
            # 按股票组织特征出现时间序列
            stock_timelines = {}
            
            for feature in features_data:
                stock_code = feature.get("stock_code", "")
                date = feature.get("date", "")
                feature_name = feature.get("pattern", feature.get("signal", ""))
                feature_type = feature.get("type", "")
                
                if not date or not stock_code:
                    continue
                
                if stock_code not in stock_timelines:
                    stock_timelines[stock_code] = {}
                
                if date not in stock_timelines[stock_code]:
                    stock_timelines[stock_code][date] = []
                
                stock_timelines[stock_code][date].append({
                    "feature": feature_name,
                    "type": feature_type
                })
            
            # 发现序列模式
            sequence_patterns = {}
            
            # 提取最长为3的序列模式
            max_sequence_length = 3
            
            for stock_code, timeline in stock_timelines.items():
                # 按日期排序
                sorted_dates = sorted(timeline.keys())
                
                # 对于每个日期，检查后续日期中的特征序列
                for i, date in enumerate(sorted_dates[:-1]):
                    current_features = timeline[date]
                    
                    # 对当前日期的每个特征
                    for current_feature in current_features:
                        # 寻找后续日期的特征序列
                        for sequence_length in range(2, min(max_sequence_length + 1, len(sorted_dates) - i)):
                            sequence = [current_feature]
                            
                            # 添加后续日期的特征
                            is_valid_sequence = True
                            for j in range(1, sequence_length):
                                next_date = sorted_dates[i + j]
                                next_features = timeline[next_date]
                                
                                if not next_features:
                                    is_valid_sequence = False
                                    break
                                
                                # 添加第一个特征
                                sequence.append(next_features[0])
                            
                            if is_valid_sequence:
                                # 生成序列标识符
                                sequence_key = " -> ".join([f"{item['type']}:{item['feature']}" for item in sequence])
                                
                                if sequence_key not in sequence_patterns:
                                    sequence_patterns[sequence_key] = {
                                        "sequence": sequence,
                                        "count": 0,
                                        "stocks": set()
                                    }
                                
                                sequence_patterns[sequence_key]["count"] += 1
                                sequence_patterns[sequence_key]["stocks"].add(stock_code)
            
            # 计算频率和支持度
            total_stocks = len(stock_timelines)
            
            for key, pattern in sequence_patterns.items():
                pattern["support"] = len(pattern["stocks"]) / total_stocks if total_stocks > 0 else 0
                pattern["stocks"] = list(pattern["stocks"])  # 转换为列表以便JSON序列化
            
            # 按支持度排序
            sorted_patterns = {k: v for k, v in sorted(
                sequence_patterns.items(), key=lambda item: item[1]["support"], reverse=True
            )}
            
            # 提取频繁序列模式
            frequent_patterns = {k: v for k, v in sorted_patterns.items() if v["support"] >= 0.1}
            
            result = {
                "total_stocks": total_stocks,
                "pattern_count": len(sequence_patterns),
                "frequent_patterns": list(frequent_patterns.values())
            }
            
            logger.info(f"特征时序模式分析完成，发现 {len(frequent_patterns)} 个频繁序列模式")
            return result
            
        except Exception as e:
            logger.error(f"分析特征时序模式时出错: {e}")
            return {}
    
    def filter_features(self, features: List[Dict[str, Any]], 
                      criteria: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        根据多维度标准筛选特征
        
        Args:
            features: 待筛选的特征列表
            criteria: 筛选标准，包含各个维度的筛选条件
            
        Returns:
            List[Dict[str, Any]]: 筛选后的特征列表
        """
        logger.info("开始多维度特征筛选")
        
        try:
            filtered_features = features.copy()
            
            # 频率筛选
            if "min_frequency" in criteria:
                min_freq = criteria["min_frequency"]
                filtered_features = [f for f in filtered_features 
                                  if f.get("frequency", 0) >= min_freq]
            
            if "max_frequency" in criteria:
                max_freq = criteria["max_frequency"]
                filtered_features = [f for f in filtered_features 
                                  if f.get("frequency", 0) <= max_freq]
            
            # 类型筛选
            if "types" in criteria and criteria["types"]:
                allowed_types = set(criteria["types"])
                filtered_features = [f for f in filtered_features 
                                  if f.get("type", "") in allowed_types]
            
            # 周期筛选
            if "periods" in criteria and criteria["periods"]:
                allowed_periods = set(criteria["periods"])
                filtered_features = [f for f in filtered_features 
                                  if not f.get("periods") or  # 如果没有周期信息，保留
                                  any(p in allowed_periods for p in f.get("periods", []))]
            
            # 关键词筛选
            if "keywords" in criteria and criteria["keywords"]:
                keywords = criteria["keywords"]
                filtered_features = [f for f in filtered_features 
                                  if any(kw.lower() in f.get("pattern", "").lower() or 
                                      kw.lower() in f.get("signal", "").lower() or
                                      kw.lower() in f.get("feature", "").lower()
                                      for kw in keywords)]
            
            # 排除关键词
            if "exclude_keywords" in criteria and criteria["exclude_keywords"]:
                exclude_keywords = criteria["exclude_keywords"]
                filtered_features = [f for f in filtered_features 
                                  if not any(kw.lower() in f.get("pattern", "").lower() or 
                                         kw.lower() in f.get("signal", "").lower() or
                                         kw.lower() in f.get("feature", "").lower()
                                         for kw in exclude_keywords)]
            
            # 最小股票数量
            if "min_stock_count" in criteria:
                min_count = criteria["min_stock_count"]
                filtered_features = [f for f in filtered_features 
                                  if f.get("count", 0) >= min_count]
            
            # 指标类型筛选
            if "indicator_types" in criteria and criteria["indicator_types"]:
                allowed_indicators = set(criteria["indicator_types"])
                filtered_features = [f for f in filtered_features 
                                  if any(ind.lower() in f.get("indicator", "").lower() or
                                      ind.lower() in f.get("feature", "").lower() or
                                      ind.lower() in f.get("pattern", "").lower()
                                      for ind in allowed_indicators)]
            
            # 自定义筛选函数
            if "custom_filter" in criteria and callable(criteria["custom_filter"]):
                custom_filter = criteria["custom_filter"]
                filtered_features = [f for f in filtered_features if custom_filter(f)]
            
            logger.info(f"特征筛选完成: 从 {len(features)} 个特征中筛选出 {len(filtered_features)} 个特征")
            return filtered_features
            
        except Exception as e:
            logger.error(f"多维度特征筛选时出错: {e}")
            return features
    
    def optimize_feature_combination(self, features: List[Dict[str, Any]], 
                                   target_metric: str,
                                   max_features: int = 10) -> Dict[str, Any]:
        """
        优化特征组合，找出最佳的特征子集
        
        Args:
            features: 候选特征列表
            target_metric: 优化目标指标，如'frequency'、'lift'等
            max_features: 最大特征数量
            
        Returns:
            Dict: 优化结果，包含最佳特征组合
        """
        logger.info(f"开始优化特征组合，目标指标: {target_metric}")
        
        try:
            # 如果特征数量不多，直接返回全部特征
            if len(features) <= max_features:
                return {
                    "optimized_features": features,
                    "feature_count": len(features),
                    "optimization_metric": target_metric,
                    "score": sum(f.get(target_metric, 0) for f in features)
                }
            
            # 根据目标指标对特征排序
            sorted_features = sorted(features, 
                                  key=lambda x: x.get(target_metric, 0), 
                                  reverse=True)
            
            # 贪婪算法选择特征
            selected_features = []
            current_score = 0
            feature_overlap = {}  # 记录特征重叠情况
            
            # 选择第一个特征
            first_feature = sorted_features[0]
            selected_features.append(first_feature)
            current_score += first_feature.get(target_metric, 0)
            
            # 使用贪婪算法选择剩余特征
            remaining_features = sorted_features[1:]
            
            while len(selected_features) < max_features and remaining_features:
                best_feature = None
                best_score_gain = 0
                
                for feature in remaining_features:
                    # 计算增加此特征的分数增益
                    score_gain = feature.get(target_metric, 0)
                    
                    # 考虑与已选特征的重叠度
                    overlap_penalty = 0
                    for selected in selected_features:
                        overlap = self._calculate_feature_overlap(feature, selected)
                        if overlap > 0:
                            overlap_penalty += overlap * 0.5  # 惩罚系数
                    
                    # 应用惩罚
                    adjusted_gain = score_gain - overlap_penalty
                    
                    if adjusted_gain > best_score_gain:
                        best_score_gain = adjusted_gain
                        best_feature = feature
                
                # 如果找不到能提供正收益的特征，则停止
                if best_feature is None or best_score_gain <= 0:
                    break
                    
                # 添加最佳特征
                selected_features.append(best_feature)
                current_score += best_score_gain
                remaining_features.remove(best_feature)
            
            # 计算组合得分
            combined_score = sum(f.get(target_metric, 0) for f in selected_features)
            
            # 构建结果
            result = {
                "optimized_features": selected_features,
                "feature_count": len(selected_features),
                "optimization_metric": target_metric,
                "score": combined_score,
                "feature_combinations": self._generate_feature_combinations(selected_features)
            }
            
            logger.info(f"特征组合优化完成，选择了 {len(selected_features)} 个特征，组合得分: {combined_score}")
            return result
            
        except Exception as e:
            logger.error(f"优化特征组合时出错: {e}")
            return {
                "optimized_features": features[:max_features],
                "feature_count": min(len(features), max_features),
                "optimization_metric": target_metric,
                "score": sum(f.get(target_metric, 0) for f in features[:max_features]),
                "error": str(e)
            }
    
    def _calculate_feature_overlap(self, feature1: Dict[str, Any], 
                                feature2: Dict[str, Any]) -> float:
        """计算两个特征之间的重叠度"""
        # 检查特征类型
        if feature1.get("type", "") != feature2.get("type", ""):
            return 0.0
        
        # 检查特征名称相似性
        name1 = feature1.get("pattern", feature1.get("signal", feature1.get("feature", "")))
        name2 = feature2.get("pattern", feature2.get("signal", feature2.get("feature", "")))
        
        if name1 == name2:
            return 1.0
        
        # 简单的相似度计算 - 可以用更复杂的算法替换
        # 检查是否有共同的关键词
        words1 = set(name1.lower().split())
        words2 = set(name2.lower().split())
        
        common_words = words1.intersection(words2)
        
        if not common_words:
            return 0.0
            
        # 计算Jaccard相似度
        similarity = len(common_words) / len(words1.union(words2))
        
        return similarity
    
    def _generate_feature_combinations(self, features: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """生成特征组合"""
        combinations = []
        
        # 生成2个特征的组合
        if len(features) >= 2:
            for i in range(len(features)):
                for j in range(i+1, len(features)):
                    combo = {
                        "features": [features[i], features[j]],
                        "combo_type": "AND",
                        "description": f"{self._get_feature_name(features[i])} AND {self._get_feature_name(features[j])}"
                    }
                    combinations.append(combo)
        
        # 生成3个特征的组合 (限制数量)
        if len(features) >= 3:
            for i in range(min(5, len(features))):
                for j in range(i+1, min(6, len(features))):
                    for k in range(j+1, min(7, len(features))):
                        combo = {
                            "features": [features[i], features[j], features[k]],
                            "combo_type": "AND",
                            "description": f"{self._get_feature_name(features[i])} AND {self._get_feature_name(features[j])} AND {self._get_feature_name(features[k])}"
                        }
                        combinations.append(combo)
        
        return combinations
    
    def _get_feature_name(self, feature: Dict[str, Any]) -> str:
        """获取特征的显示名称"""
        return feature.get("pattern", feature.get("signal", feature.get("feature", "未知特征")))